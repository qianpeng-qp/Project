from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class do_data:

    @staticmethod
    def do_xls(filepath, sheetname):
        workbook = load_workbook(filepath)
        ws: Worksheet = workbook[sheetname]
        test_data = []

        for row in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
            test_data.append(row)
        return test_data
